import os
import time
import shutil
import pandas as pd
import win32com.client  # Biblioteca para controlar o Excel oficialmente
from playwright.sync_api import sync_playwright
from datetime import datetime
from openpyxl import load_workbook
from inputimeout import inputimeout, TimeoutOccurred

# ================= CONFIGURAÇÕES DE TESTE =================
USER = "seu_usuario@dominio.com"
PASSWORD = "sua_password_segura"
URL_LOGIN = "https://desk.auvo.com.br/Login"

# Seus caminhos de teste no OneDrive
BASE_DIR = r"C:\Caminho\Para\Seu\Diretorio\Sincronizado"
NOME_ARQUIVO_LOCAL = os.path.join(BASE_DIR, "Relatório Geral.xlsx")
PASTA_DOWNLOADS = os.path.join(BASE_DIR, "downloads_temporarios")
ARQUIVO_LOG = os.path.join(BASE_DIR, "status_robo.txt")
NOME_ABA = "Relatório"
CAMINHO_SHAREPOINT = r"C:\Caminho\Para\Seu\Diretorio\Sincronizado"

periodos = [
    {"inicio": "2025-01-01", "fim": "2025-03-31", "nome": "Tri1-2025"},
    {"inicio": "2025-04-01", "fim": "2025-06-30", "nome": "Tri2-2025"},
    {"inicio": "2025-07-01", "fim": "2025-09-30", "nome": "Tri3-2025"},
    {"inicio": "2025-10-01", "fim": "2025-12-31", "nome": "Tri4-2025"},
    {"inicio": "2026-01-01", "fim": "2026-03-31", "nome": "Tri1-2026"},
    {"inicio": "2026-04-01", "fim": "2026-06-30", "nome": "Tri2-2026"},
    {"inicio": "2026-07-01", "fim": "2026-09-30", "nome": "Tri3-2026"},
    {"inicio": "2026-10-01", "fim": "2026-12-31", "nome": "Tri4-2026"},
]

def escrever_log(mensagem):
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    try:
        with open(ARQUIVO_LOG, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {mensagem}\n")
        print(mensagem)
    except: pass

def forcar_recalculo_excel(caminho_arquivo):
    """Abre o Excel invisivelmente para forçar o cálculo das fórmulas e validar o arquivo."""
    try:
        escrever_log("⚙️ Forçando recálculo oficial das fórmulas no Excel...")
        # Inicializa o aplicativo Excel
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Roda em segundo plano
        excel.DisplayAlerts = False # Ignora avisos chatos
        
        # Abre, salva e fecha
        wb = excel.Workbooks.Open(caminho_arquivo)
        wb.Save()
        wb.Close()
        excel.Quit()
        escrever_log("✅ Recálculo concluído. O Power BI agora verá os dados atualizados!")
    except Exception as e:
        escrever_log(f"⚠️ Erro ao forçar recálculo: {e}")

def deve_baixar_novamente():
    if not os.path.exists(PASTA_DOWNLOADS):
        os.makedirs(PASTA_DOWNLOADS)
        return True
    
    arquivos_existentes = [f for f in os.listdir(PASTA_DOWNLOADS) if f.endswith('.xlsx')]
    
    if arquivos_existentes:
        print(f"\n📂 Encontrei {len(arquivos_existentes)} planilhas locais.")
        try:
            resposta = inputimeout(prompt='❓ Usar arquivos existentes? (s/n) [Padrão "n" em 10s]: ', timeout=10).lower()
            if resposta == 's':
                escrever_log("⏩ Pulando download e usando arquivos locais.")
                return False
        except TimeoutOccurred:
            print("\n⏰ Tempo esgotado. Iniciando novo download...")
    
    escrever_log("🧹 Limpando pasta temporária...")
    for arq in os.listdir(PASTA_DOWNLOADS):
        try: os.remove(os.path.join(PASTA_DOWNLOADS, arq))
        except: pass
    return True

def automatizar_download():
    with sync_playwright() as p:
        escrever_log("🚀 Iniciando navegador...")
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        try:
            page.goto(URL_LOGIN)
            page.fill("#Login", USER)
            page.fill("#Senha", PASSWORD)
            page.click(".btnLogin")
            page.wait_for_load_state("networkidle")
            time.sleep(5)

            for p_info in periodos:
                escrever_log(f"🏠 Baixando: {p_info['nome']}")
                page.click('img[alt="homepage"]')
                time.sleep(3)
                page.wait_for_selector("#MostrarFiltro")
                page.click("#MostrarFiltro")
                page.fill("#DataInicioFiltro", p_info['inicio'])
                page.fill("#DataFimFiltro", p_info['fim'])
                
                while True:
                    remover = page.query_selector(".select2-selection__choice__remove")
                    if remover and remover.is_visible():
                        remover.click()
                        time.sleep(0.2)
                    else: break
                
                page.click("#btnAplicarFiltro")
                time.sleep(4)
                page.click("#btnExporteExcel")
                
                try:
                    page.wait_for_selector(".swal2-confirm", state="visible", timeout=10000)
                    with page.expect_download(timeout=60000) as download_info:
                        page.click(".swal2-confirm")
                    download = download_info.value
                    download.save_as(os.path.join(PASTA_DOWNLOADS, f"{p_info['nome']}.xlsx"))
                except: escrever_log(f"⏭️ Sem dados em {p_info['nome']}")
            
            browser.close()
            return True
        except Exception as e:
            escrever_log(f"❌ Erro no download: {e}")
            browser.close()
            return False

def consolidar_preservando_formulas():
    escrever_log("📊 Consolidando dados (A até AH)...")
    lista_dfs = []
    
    for p_info in periodos:
        caminho = os.path.join(PASTA_DOWNLOADS, f"{p_info['nome']}.xlsx")
        if os.path.exists(caminho):
            try:
                df = pd.read_excel(caminho).iloc[:, 0:34]
                if not df.empty:
                    lista_dfs.append(df)
            except: pass

    if not lista_dfs:
        escrever_log("🛑 INTERROMPIDO: Nenhum dado novo encontrado.")
        return False

    try:
        df_final = pd.concat(lista_dfs, ignore_index=True)
        wb = load_workbook(NOME_ARQUIVO_LOCAL)
        ws = wb[NOME_ABA]

        escrever_log("🧹 Limpando dados antigos (A-AH)...")
        for row in ws.iter_rows(min_row=2, max_col=34, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        escrever_log(f"🖊️ Inserindo {len(df_final)} novas linhas...")
        for i, row in enumerate(df_final.values.tolist()):
            for j, value in enumerate(row):
                ws.cell(row=i+2, column=j+1, value=value)

        wb.save(NOME_ARQUIVO_LOCAL)
        
        # --- FORÇA O RECÁLCULO ANTES DA CÓPIA ---
        forcar_recalculo_excel(NOME_ARQUIVO_LOCAL)
        
        # Sincroniza se o caminho for diferente
        if NOME_ARQUIVO_LOCAL.lower() != CAMINHO_SHAREPOINT.lower():
            shutil.copy2(NOME_ARQUIVO_LOCAL, CAMINHO_SHAREPOINT)
            escrever_log("🚀 Arquivo sincronizado com o destino final!")
        
        return True
    except Exception as e:
        escrever_log(f"❌ Erro na consolidação: {e}")
        return False

if __name__ == "__main__":
    os.environ["NODE_TLS_REJECT_UNAUTHORIZED"] = "0"
    
    if deve_baixar_novamente():
        if not automatizar_download():
            escrever_log("⚠️ Falha no download. Encerrando.")
            exit()
    
    if consolidar_preservando_formulas():
        escrever_log("🏁 Processo concluído com sucesso.")